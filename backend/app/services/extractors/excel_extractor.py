from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime, time
from itertools import zip_longest
from pathlib import Path
from typing import Any, Iterable

import xlrd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
GROUPED_ROW_WINDOW = 10
GROUPED_ROW_OVERLAP = 2

COMMON_HEADER_WORDS = {
    "account",
    "amount",
    "application",
    "author",
    "code",
    "column",
    "date",
    "description",
    "field",
    "id",
    "name",
    "number",
    "owner",
    "record",
    "records",
    "reference",
    "source",
    "status",
    "target",
    "task",
    "type",
    "value",
    "version",
}


@dataclass
class ExcelRowRecord:
    row_number: int
    values: dict[str, Any]
    text: str
    row_type: str = "data"
    section_title: str | None = None
    header_row_number: int | None = None
    formulas: dict[str, str] | None = None
    cell_references: dict[str, str] | None = None


@dataclass
class ExcelSheetRecord:
    sheet_name: str
    state: str
    max_row: int
    max_column: int
    headers: list[str]
    row_count: int
    preview_rows: list[ExcelRowRecord]
    data_row_count: int = 0
    formula_cell_count: int = 0
    header_row_numbers: list[int] | None = None
    extraction_warnings: list[str] | None = None


@dataclass
class ExcelWorkbookRecord:
    file_path: str
    file_name: str
    file_type: str
    sheet_count: int
    sheets: list[ExcelSheetRecord]


@dataclass
class ExcelIngestionRecord:
    record_id: str
    text: str
    metadata: dict[str, Any]


@dataclass
class _CellRecord:
    row_number: int
    column_number: int
    coordinate: str
    value: Any
    formula: str | None = None


def _clean_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _cell_value_to_text(value: Any) -> str:
    cleaned = _clean_cell_value(value)
    if cleaned is None:
        return ""
    if isinstance(cleaned, datetime):
        return cleaned.isoformat(sep=" ")
    if isinstance(cleaned, date):
        return cleaned.isoformat()
    if isinstance(cleaned, time):
        return cleaned.isoformat()
    return str(cleaned)


def _normalize_header(value: Any, idx: int) -> str:
    cleaned = _cell_value_to_text(value)
    return cleaned or f"column_{idx}"


def _normalize_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values, start=1):
        header = _normalize_header(value, idx)
        count = seen.get(header, 0) + 1
        seen[header] = count
        headers.append(header if count == 1 else f"{header}_{count}")
    return headers


def _is_placeholder_header(header: str) -> bool:
    return header.startswith("column_")


def _prune_row_map(row_map: dict[str, Any]) -> dict[str, Any]:
    cleaned_map: dict[str, Any] = {}
    for key, value in row_map.items():
        cleaned_value = _clean_cell_value(value)
        if cleaned_value is None:
            continue
        cleaned_map[key] = cleaned_value
    return cleaned_map


def _is_likely_template_value(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    return text.startswith("*") or text.startswith("#")


def _filter_meaningful_row_map(row_map: dict[str, Any]) -> dict[str, Any]:
    filtered: dict[str, Any] = {}
    for key, value in row_map.items():
        if value is None:
            continue
        if _is_placeholder_header(key) and _is_likely_template_value(value):
            continue
        filtered[key] = value
    return filtered


def _should_skip_row(row_map: dict[str, Any]) -> bool:
    if not row_map:
        return True

    values = list(row_map.values())
    non_empty_count = len(values)
    if non_empty_count == 0:
        return True

    template_like_count = sum(1 for value in values if _is_likely_template_value(value))
    placeholder_header_count = sum(1 for header in row_map if _is_placeholder_header(header))

    if non_empty_count >= 3 and template_like_count / non_empty_count >= 0.6:
        return True

    if placeholder_header_count == len(row_map) and template_like_count > 0:
        return True

    return False


def _dedupe_preserve_order(values: Iterable[Any]) -> list[Any]:
    seen: set[str] = set()
    output: list[Any] = []
    for value in values:
        key = str(value)
        if key in seen:
            continue
        seen.add(key)
        output.append(value)
    return output


def _meaningful_headers(headers: list[str], row_maps: list[dict[str, Any]]) -> list[str]:
    if not row_maps:
        return _dedupe_preserve_order(header for header in headers if not _is_placeholder_header(header))

    value_keys = {key for row_map in row_maps for key in row_map.keys()}
    return [
        header
        for header in _dedupe_preserve_order(headers)
        if header in value_keys or not _is_placeholder_header(header)
    ]


def _row_to_text(row_map: dict[str, Any], formulas: dict[str, str] | None = None) -> str:
    parts = [f"{key}: {_cell_value_to_text(value)}" for key, value in row_map.items()]
    if formulas:
        parts.extend(f"{key} formula: {formula}" for key, formula in formulas.items())
    return ", ".join(part for part in parts if part)


def _header_word_score(text: str) -> bool:
    normalized = "".join(ch if ch.isalnum() else " " for ch in text.lower())
    words = [word for word in normalized.split() if word]
    return any(word in COMMON_HEADER_WORDS for word in words)


def _looks_like_header_row(values: list[Any]) -> bool:
    non_empty = [_cell_value_to_text(value) for value in values if _clean_cell_value(value) is not None]
    if len(non_empty) < 2:
        return False

    marker_count = sum(1 for value in non_empty if value.startswith(("*", "#")))
    common_header_count = sum(1 for value in non_empty if _header_word_score(value))
    short_text_count = sum(1 for value in non_empty if len(value) <= 80)
    numeric_like_count = sum(1 for value in non_empty if isinstance(value, (int, float)))

    if numeric_like_count:
        return False
    if marker_count:
        return True
    return common_header_count / len(non_empty) > 0.55 and short_text_count / len(non_empty) >= 0.8


def _section_title_from_row(values: list[Any]) -> str | None:
    non_empty = [_cell_value_to_text(value) for value in values if _clean_cell_value(value) is not None]
    if len(non_empty) != 1:
        return None
    value = non_empty[0]
    return value if value and len(value) <= 160 else None


def _pad_headers(headers: list[str], length: int) -> list[str]:
    padded = headers[:]
    if len(padded) < length:
        padded.extend(f"column_{idx}" for idx in range(len(padded) + 1, length + 1))
    return padded


def _build_row_payload(
    row_number: int,
    cells: list[_CellRecord],
    headers: list[str],
) -> tuple[dict[str, Any], dict[str, str], dict[str, str]]:
    padded_headers = _pad_headers(headers, len(cells))
    values: dict[str, Any] = {}
    formulas: dict[str, str] = {}
    cell_references: dict[str, str] = {}

    for idx, cell in enumerate(cells):
        header = padded_headers[idx]
        value = _clean_cell_value(cell.value)
        if value is None and cell.formula:
            value = cell.formula
        if value is None:
            continue
        values[header] = value
        cell_references[header] = cell.coordinate
        if cell.formula:
            formulas[header] = cell.formula

    return values, formulas, cell_references


def _openpyxl_cell_record(data_cell: Cell | None, formula_cell: Cell | None, row_number: int, column_number: int) -> _CellRecord:
    coordinate = (
        getattr(data_cell, "coordinate", None)
        or getattr(formula_cell, "coordinate", None)
        or f"{get_column_letter(column_number)}{row_number}"
    )
    formula = None
    formula_value = getattr(formula_cell, "value", None)
    if isinstance(formula_value, str) and formula_value.startswith("="):
        formula = formula_value
    value = getattr(data_cell, "value", None)
    return _CellRecord(row_number=row_number, column_number=column_number, coordinate=coordinate, value=value, formula=formula)


def _iter_openpyxl_rows(data_sheet, formula_sheet) -> Iterable[tuple[int, list[_CellRecord]]]:
    data_rows = data_sheet.iter_rows()
    formula_rows = formula_sheet.iter_rows()
    for row_number, (data_row, formula_row) in enumerate(zip_longest(data_rows, formula_rows, fillvalue=()), start=1):
        max_columns = max(len(data_row), len(formula_row))
        cells = [
            _openpyxl_cell_record(
                data_row[idx] if idx < len(data_row) else None,
                formula_row[idx] if idx < len(formula_row) else None,
                row_number,
                idx + 1,
            )
            for idx in range(max_columns)
        ]
        yield row_number, cells


def _xls_cell_value(workbook, worksheet, row_idx: int, col_idx: int) -> Any:
    cell = worksheet.cell(row_idx, col_idx)
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.biffh.error_text_from_code.get(cell.value, f"#ERROR({cell.value})")
    return cell.value


def _iter_xlrd_rows(workbook, worksheet) -> Iterable[tuple[int, list[_CellRecord]]]:
    for row_idx in range(worksheet.nrows):
        row_number = row_idx + 1
        cells = [
            _CellRecord(
                row_number=row_number,
                column_number=col_idx + 1,
                coordinate=f"{get_column_letter(col_idx + 1)}{row_number}",
                value=_xls_cell_value(workbook, worksheet, row_idx, col_idx),
            )
            for col_idx in range(worksheet.ncols)
        ]
        yield row_number, cells


def _extract_sheet(
    *,
    sheet_name: str,
    state: str,
    max_row: int,
    max_column: int,
    rows: Iterable[tuple[int, list[_CellRecord]]],
    preview_limit: int | None,
) -> ExcelSheetRecord:
    current_headers: list[str] = []
    current_header_row_number: int | None = None
    header_row_numbers: list[int] = []
    all_headers: list[str] = []
    retained_row_maps: list[dict[str, Any]] = []
    preview_rows: list[ExcelRowRecord] = []
    section_title: str | None = None
    formula_cell_count = 0
    extraction_warnings: list[str] = []

    for row_number, cells in rows:
        cleaned_values = [_clean_cell_value(cell.value) for cell in cells]
        row_formula_count = sum(1 for cell in cells if cell.formula)
        formula_cell_count += row_formula_count

        if not any(value is not None or cell.formula for value, cell in zip(cleaned_values, cells)):
            continue

        header_candidate_values = [cell.formula if value is None and cell.formula else value for value, cell in zip(cleaned_values, cells)]
        section_candidate = _section_title_from_row(header_candidate_values)
        if section_candidate:
            section_title = section_candidate

        is_header = _looks_like_header_row(header_candidate_values)
        if is_header:
            current_headers = _normalize_headers(header_candidate_values)
            current_header_row_number = row_number
            header_row_numbers.append(row_number)
            all_headers.extend(current_headers)

        headers = current_headers or _normalize_headers([None] * len(cells))
        row_map, formulas, cell_references = _build_row_payload(row_number, cells, headers)
        row_map = _prune_row_map(row_map)

        row_type = "header" if is_header else "section" if section_candidate else "data"
        filtered_row_map = row_map if row_type != "data" else _filter_meaningful_row_map(row_map)
        if row_type == "data" and _should_skip_row(filtered_row_map):
            continue

        text = _row_to_text(filtered_row_map, formulas)
        if not text:
            continue

        retained_row_maps.append(filtered_row_map)
        preview_rows.append(
            ExcelRowRecord(
                row_number=row_number,
                values=filtered_row_map,
                text=text,
                row_type=row_type,
                section_title=section_title,
                header_row_number=current_header_row_number,
                formulas=formulas or None,
                cell_references=cell_references or None,
            )
        )
        if preview_limit is not None and len(preview_rows) >= preview_limit:
            extraction_warnings.append(f"Sheet extraction stopped at preview_limit={preview_limit}")
            break

    headers = _meaningful_headers(all_headers or current_headers, retained_row_maps)
    data_row_count = sum(1 for row in preview_rows if row.row_type == "data")

    return ExcelSheetRecord(
        sheet_name=sheet_name,
        state=state,
        max_row=max_row,
        max_column=max_column,
        headers=headers,
        row_count=len(preview_rows),
        preview_rows=preview_rows,
        data_row_count=data_row_count,
        formula_cell_count=formula_cell_count,
        header_row_numbers=header_row_numbers or None,
        extraction_warnings=extraction_warnings or None,
    )


def _extract_openpyxl_workbook(path: Path, preview_limit: int | None, keep_vba: bool) -> list[ExcelSheetRecord]:
    data_workbook = load_workbook(filename=path, data_only=True, keep_vba=keep_vba, read_only=True)
    formula_workbook = load_workbook(filename=path, data_only=False, keep_vba=keep_vba, read_only=True)
    try:
        sheets: list[ExcelSheetRecord] = []
        for data_sheet in data_workbook.worksheets:
            formula_sheet = formula_workbook[data_sheet.title]
            sheets.append(
                _extract_sheet(
                    sheet_name=data_sheet.title,
                    state=data_sheet.sheet_state,
                    max_row=data_sheet.max_row or 0,
                    max_column=data_sheet.max_column or 0,
                    rows=_iter_openpyxl_rows(data_sheet, formula_sheet),
                    preview_limit=preview_limit,
                )
            )
        return sheets
    finally:
        _close_openpyxl_workbook(data_workbook)
        _close_openpyxl_workbook(formula_workbook)


def _close_openpyxl_workbook(workbook) -> None:
    workbook.close()
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        vba_archive.close()


def _extract_xls_workbook(path: Path, preview_limit: int | None) -> list[ExcelSheetRecord]:
    workbook = xlrd.open_workbook(str(path), formatting_info=False)
    try:
        sheets: list[ExcelSheetRecord] = []
        for worksheet in workbook.sheets():
            sheets.append(
                _extract_sheet(
                    sheet_name=worksheet.name,
                    state="visible",
                    max_row=worksheet.nrows,
                    max_column=worksheet.ncols,
                    rows=_iter_xlrd_rows(workbook, worksheet),
                    preview_limit=preview_limit,
                )
            )
        return sheets
    finally:
        workbook.release_resources()


def extract_excel_workbook(file_path: str | Path, preview_limit: int | None = None) -> ExcelWorkbookRecord:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix not in SUPPORTED_EXCEL_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXCEL_EXTENSIONS))
        raise ValueError(f"Unsupported Excel file type: {suffix}. Supported: {supported}")

    if suffix in {".xlsx", ".xlsm"}:
        sheets = _extract_openpyxl_workbook(path, preview_limit, keep_vba=(suffix == ".xlsm"))
    else:
        sheets = _extract_xls_workbook(path, preview_limit)

    return ExcelWorkbookRecord(
        file_path=str(path.resolve()),
        file_name=path.name,
        file_type=suffix.lstrip("."),
        sheet_count=len(sheets),
        sheets=sheets,
    )


def workbook_record_to_dict(record: ExcelWorkbookRecord) -> dict[str, Any]:
    return asdict(record)


def workbook_record_to_ingestion_records(record: ExcelWorkbookRecord) -> list[ExcelIngestionRecord]:
    ingestion_records: list[ExcelIngestionRecord] = []

    for sheet in record.sheets:
        step = max(1, GROUPED_ROW_WINDOW - GROUPED_ROW_OVERLAP)
        for start in range(0, len(sheet.preview_rows), step):
            row_group = sheet.preview_rows[start : start + GROUPED_ROW_WINDOW]
            if not row_group:
                continue
            grouped_values: list[dict[str, Any]] = []
            grouped_lines: list[str] = []
            row_numbers: list[int] = []
            row_types: list[str] = []
            section_titles: list[str] = []
            formulas: dict[str, dict[str, str]] = {}
            cell_references: dict[str, dict[str, str]] = {}

            for row in row_group:
                non_empty_values = {
                    key: value
                    for key, value in row.values.items()
                    if value is not None and not (_is_placeholder_header(key) and _is_likely_template_value(value))
                }
                if not non_empty_values:
                    continue

                grouped_values.append(non_empty_values)
                row_numbers.append(row.row_number)
                if row.row_type not in row_types:
                    row_types.append(row.row_type)
                if row.section_title and row.section_title not in section_titles:
                    section_titles.append(row.section_title)
                if row.formulas:
                    formulas[str(row.row_number)] = row.formulas
                if row.cell_references:
                    cell_references[str(row.row_number)] = row.cell_references

                prefix = f"Row {row.row_number}"
                if row.row_type != "data":
                    prefix = f"{prefix} ({row.row_type})"
                if row.section_title and row.row_type == "data":
                    prefix = f"{prefix} [{row.section_title}]"
                grouped_lines.append(f"{prefix}: {row.text}")

            if not grouped_values:
                continue

            value_keys: list[str] = []
            for value_map in grouped_values:
                for key in value_map.keys():
                    if key not in value_keys:
                        value_keys.append(key)

            non_placeholder_headers = [
                header for header in sheet.headers if header in value_keys or not _is_placeholder_header(header)
            ]

            heading = f"Sheet: {sheet.sheet_name}"
            if section_titles:
                heading = f"{heading} | Section: {section_titles[0]}"

            metadata = {
                "document_type": "excel",
                "file_name": record.file_name,
                "file_path": record.file_path,
                "file_type": record.file_type,
                "sheet_name": sheet.sheet_name,
                "sheet_state": sheet.state,
                "row_number": row_numbers[0],
                "row_start": row_numbers[0],
                "row_end": row_numbers[-1],
                "row_numbers": row_numbers,
                "row_types": row_types,
                "section_titles": section_titles,
                "headers": non_placeholder_headers,
                "header_row_numbers": sheet.header_row_numbers or [],
                "values": grouped_values,
                "value_keys": value_keys,
                "formulas": formulas,
                "cell_references": cell_references,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "data_row_count": sheet.data_row_count,
                "formula_cell_count": sheet.formula_cell_count,
                "chunk_type": "excel_row_group",
            }
            ingestion_records.append(
                ExcelIngestionRecord(
                    record_id=f"{record.file_name}:{sheet.sheet_name}:{row_numbers[0]}-{row_numbers[-1]}",
                    text=(
                        f"Workbook: {record.file_name} | "
                        f"Sheet: {sheet.sheet_name} | "
                        f"Rows: {row_numbers[0]}-{row_numbers[-1]} | "
                        + " || ".join(grouped_lines)
                    ),
                    metadata=metadata,
                )
            )

    return ingestion_records


def ingestion_records_to_dict(records: list[ExcelIngestionRecord]) -> list[dict[str, Any]]:
    return [asdict(record) for record in records]


def workbook_record_to_text(record: ExcelWorkbookRecord) -> str:
    lines = [
        f"Workbook: {record.file_name}",
        f"Path: {record.file_path}",
        f"Type: {record.file_type}",
        f"Sheets: {record.sheet_count}",
    ]

    for sheet in record.sheets:
        lines.append("")
        lines.append(f"## Sheet: {sheet.sheet_name}")
        lines.append(f"State: {sheet.state}")
        lines.append(f"Dimensions: max_row={sheet.max_row}, max_column={sheet.max_column}")
        lines.append(f"Headers: {', '.join(sheet.headers) if sheet.headers else '[none detected]'}")
        lines.append(f"Header rows: {', '.join(map(str, sheet.header_row_numbers or [])) or '[none detected]'}")
        lines.append(f"Extracted row count: {sheet.row_count}")
        lines.append(f"Data row count: {sheet.data_row_count}")
        lines.append(f"Formula cell count: {sheet.formula_cell_count}")
        if sheet.extraction_warnings:
            lines.append(f"Warnings: {'; '.join(sheet.extraction_warnings)}")

        for row in sheet.preview_rows:
            label = f"- Row {row.row_number}"
            if row.row_type != "data":
                label = f"{label} ({row.row_type})"
            if row.section_title and row.row_type == "data":
                label = f"{label} [{row.section_title}]"
            lines.append(f"{label}: {row.text}")

    return "\n".join(lines)


def ingestion_records_to_text(records: list[ExcelIngestionRecord]) -> str:
    lines = [f"Ingestion records: {len(records)}"]

    for record in records:
        metadata = record.metadata
        lines.append("")
        lines.append(f"## Record ID: {record.record_id}")
        lines.append(f"LLM Text: {record.text}")
        lines.append(
            "Metadata: "
            f"sheet_name={metadata.get('sheet_name')}, "
            f"rows={metadata.get('row_start')}-{metadata.get('row_end')}, "
            f"file_name={metadata.get('file_name')}, "
            f"sections={metadata.get('section_titles')}"
        )
        lines.append(f"Values: {metadata.get('values')}")

    return "\n".join(lines)
